#!/usr/bin/env python3
"""
Markdown测试用例转Excel转换器

该脚本用于扫描指定目录下的Markdown文件，解析其中的测试用例表格，
并将所有测试用例整合到一个Excel文件中导出。

作者: AI Assistant
创建时间: 2025-09-05
"""

import os
import re
import sys
from pathlib import Path
from typing import List, Dict, Optional, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class MarkdownToExcelConverter:
    """Markdown测试用例转Excel转换器类"""
    
    def __init__(self, input_dir: str, output_file: str = "测试用例汇总.xlsx"):
        """
        初始化转换器
        
        Args:
            input_dir: 输入目录路径
            output_file: 输出Excel文件名
        """
        self.input_dir = Path(input_dir)
        self.output_file = output_file
        self.test_cases: List[Dict[str, str]] = []
        
        # 定义Excel列标题
        self.columns = [
            "文件名",
            "覆盖范围", 
            "测试用例ID",
            "测试用例名称",
            "前置条件",
            "测试步骤",
            "预期结果",
            "实际结果",
            "测试状态"
        ]
    
    def scan_markdown_files(self) -> List[Path]:
        """
        递归扫描目录下的所有Markdown文件
        
        Returns:
            Markdown文件路径列表
        """
        if not self.input_dir.exists():
            raise FileNotFoundError(f"目录不存在: {self.input_dir}")
        
        md_files = []
        for root, dirs, files in os.walk(self.input_dir):
            for file in files:
                if file.endswith('.md'):
                    md_files.append(Path(root) / file)
        
        print(f"扫描到 {len(md_files)} 个Markdown文件")
        return md_files
    
    def extract_coverage_info(self, content: str) -> str:
        """
        提取覆盖范围信息
        
        Args:
            content: 文件内容
            
        Returns:
            覆盖范围文本
        """
        # 匹配 ## 覆盖范围 部分
        pattern = r'## 覆盖范围\s*\n(.*?)(?=\n##|\n$|$)'
        match = re.search(pattern, content, re.DOTALL)
        
        if match:
            coverage_text = match.group(1).strip()
            # 移除markdown列表符号
            coverage_text = re.sub(r'^- ', '', coverage_text, flags=re.MULTILINE)
            return coverage_text
        
        return ""
    
    def extract_test_cases(self, content: str) -> List[Dict[str, str]]:
        """
        提取测试用例表格数据
        
        Args:
            content: 文件内容
            
        Returns:
            测试用例列表
        """
        test_cases = []
        
        # 查找表格部分
        table_pattern = r'\| 测试用例ID.*?\n\|---.*?\n((?:\|.*?\n)*)'
        table_match = re.search(table_pattern, content, re.DOTALL)
        
        if not table_match:
            return test_cases
        
        table_content = table_match.group(1)
        
        # 解析表格行
        for line in table_content.strip().split('\n'):
            if line.strip() and line.startswith('|'):
                # 分割表格列
                columns = [col.strip() for col in line.split('|')[1:-1]]
                
                if len(columns) >= 7:  # 确保有足够的列
                    test_case = {
                        "测试用例ID": columns[0],
                        "测试用例名称": columns[1], 
                        "前置条件": columns[2],
                        "测试步骤": columns[3],
                        "预期结果": columns[4],
                        "实际结果": columns[5],
                        "测试状态": columns[6]
                    }
                    test_cases.append(test_case)
        
        return test_cases
    
    def parse_markdown_file(self, file_path: Path) -> Tuple[str, List[Dict[str, str]]]:
        """
        解析单个Markdown文件
        
        Args:
            file_path: 文件路径
            
        Returns:
            (覆盖范围, 测试用例列表)
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            coverage = self.extract_coverage_info(content)
            test_cases = self.extract_test_cases(content)
            
            return coverage, test_cases
            
        except Exception as e:
            print(f"解析文件失败 {file_path}: {e}")
            return "", []
    
    def process_all_files(self) -> None:
        """处理所有Markdown文件"""
        md_files = self.scan_markdown_files()
        
        total_cases = 0
        for i, file_path in enumerate(md_files, 1):
            print(f"正在处理 ({i}/{len(md_files)}): {file_path.name}")
            
            coverage, test_cases = self.parse_markdown_file(file_path)
            
            # 为每个测试用例添加文件信息
            for test_case in test_cases:
                test_case["文件名"] = file_path.name
                test_case["覆盖范围"] = coverage
                self.test_cases.append(test_case)
            
            total_cases += len(test_cases)
            print(f"  提取到 {len(test_cases)} 个测试用例")
        
        print(f"\n总计提取到 {total_cases} 个测试用例")
    
    def create_excel_file(self) -> None:
        """创建Excel文件"""
        if not self.test_cases:
            print("没有找到测试用例数据")
            return
        
        # 创建DataFrame
        df = pd.DataFrame(self.test_cases, columns=self.columns)
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例汇总"
        
        # 写入数据
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 设置样式
        self._format_excel(ws)
        
        # 保存文件
        wb.save(self.output_file)
        print(f"Excel文件已保存: {self.output_file}")
    
    def _format_excel(self, worksheet) -> None:
        """格式化Excel工作表"""
        # 设置标题行样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置列宽
        column_widths = {
            'A': 25,  # 文件名
            'B': 40,  # 覆盖范围
            'C': 15,  # 测试用例ID
            'D': 30,  # 测试用例名称
            'E': 20,  # 前置条件
            'F': 50,  # 测试步骤
            'G': 40,  # 预期结果
            'H': 50,  # 实际结果
            'I': 12   # 测试状态
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 设置行高和文本换行
        for row in worksheet.iter_rows(min_row=2):
            worksheet.row_dimensions[row[0].row].height = 60
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    def run(self) -> None:
        """运行转换器"""
        try:
            print("开始处理Markdown文件...")
            self.process_all_files()
            
            print("正在创建Excel文件...")
            self.create_excel_file()
            
            print("转换完成!")
            
        except Exception as e:
            print(f"转换过程中发生错误: {e}")
            sys.exit(1)


def main():
    """主函数"""
    # 默认配置
    input_directory = "/Users/sniperz/dockerfiles/dify-plus/docs/测试用例/"
    output_filename = "测试用例汇总.xlsx"
    
    # 检查命令行参数
    if len(sys.argv) > 1:
        input_directory = sys.argv[1]
    if len(sys.argv) > 2:
        output_filename = sys.argv[2]
    
    print(f"输入目录: {input_directory}")
    print(f"输出文件: {output_filename}")
    print("-" * 50)
    
    # 创建转换器并运行
    converter = MarkdownToExcelConverter(input_directory, output_filename)
    converter.run()


if __name__ == "__main__":
    main()
